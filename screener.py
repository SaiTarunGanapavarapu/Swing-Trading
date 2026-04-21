#!/usr/bin/env python3
"""
Indian Stock Swing Trading Screener
====================================
Screens NSE stocks through Buffett, Lynch & Graham rules.
Outputs a scored, ranked list of swing trade candidates.

Requirements:
    pip install yfinance pandas openpyxl tabulate

Usage:
    python screener.py                    # Run with default Nifty 200
    python screener.py --symbols TCS.NS INFY.NS
    python screener.py --csv stocks.csv   # CSV with 'Symbol' column
    python screener.py --index "NIFTY 50"
"""

import argparse
import math
import re
import time
import warnings
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
import requests
import yfinance as yf

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────

RISK_FREE_RATE = 7.0  # India 10Y govt bond yield (%)
SLEEP_BETWEEN_FETCHES = 1.2  # seconds, to avoid rate limiting
MAX_WORKERS = 1  # sequential to avoid yfinance bans
USE_EXTERNAL_FALLBACK = False  # Prefer yfinance-only derived metrics for consistency and speed
MONEYCONTROL_SUGGEST_URL = "https://www.moneycontrol.com/mccode/common/autosuggestion_solr.php"
MONEYCONTROL_TIMEOUT_SECONDS = 20
SCREENER_MIN_INTERVAL_SECONDS = 1.0
SCREENER_TIMEOUT_SECONDS = 20

moneycontrolSession = requests.Session()
moneycontrolRatiosCache = {}
moneycontrolFieldLabels = {
    "returnOnEquity": ["Return on Equity / Networth (%)", "Return on Networth / Equity (%)"],
    "currentRatio": ["Current Ratio (X)"],
    "enterpriseToEbitda": ["EV/EBITDA (X)"],
    "debtToEquity": ["Total Debt/Equity (X)", "Debt-Equity Ratio"],
}
screenerSession = requests.Session()
screenerTopRatiosCache = {}
screenerLastRequestTs = 0.0

# Default stock list (Nifty 50 components — extend as needed)
DEFAULT_SYMBOLS = [
    "ADANIENT.NS", "ADANIPORTS.NS", "APOLLOHOSP.NS", "ASIANPAINT.NS", "AXISBANK.NS",
    "BAJAJ-AUTO.NS", "BAJFINANCE.NS", "BAJAJFINSV.NS", "BEL.NS", "BHARTIARTL.NS",
    "CIPLA.NS", "COALINDIA.NS", "DRREDDY.NS", "EICHERMOT.NS", "ETERNAL.NS",
    "GRASIM.NS", "HCLTECH.NS", "HDFCBANK.NS", "HDFCLIFE.NS", "HINDALCO.NS",
    "HINDUNILVR.NS", "ICICIBANK.NS", "INDIGO.NS", "INFY.NS", "ITC.NS",
    "JIOFIN.NS", "JSWSTEEL.NS", "KOTAKBANK.NS", "LT.NS", "M&M.NS",
    "MARUTI.NS", "MAXHEALTH.NS", "NESTLEIND.NS", "NTPC.NS", "ONGC.NS",
    "POWERGRID.NS", "RELIANCE.NS", "SBILIFE.NS", "SHRIRAMFIN.NS", "SBIN.NS",
    "SUNPHARMA.NS", "TCS.NS", "TATACONSUM.NS", "TMPV.NS", "TATASTEEL.NS",
    "TECHM.NS", "TITAN.NS", "TRENT.NS", "ULTRACEMCO.NS", "WIPRO.NS"
]

# BankNifty components (for financial sector focus)
# DEFAULT_SYMBOLS = [
#     "HDFCBANK.NS", "ICICIBANK.NS", "AXISBANK.NS", "KOTAKBANK.NS",
#     "SBIN.NS", "INDUSINDBK.NS", "AUBANK.NS", "BANDHANBNK.NS",
#     "FEDERALBNK.NS", "IDFCFIRSTB.NS", "PNB.NS", "BANKBARODA.NS"
# ]

# def load_default_symbols(csv_path: str = "EQUITY_L.csv") -> list[str]:
#     """Load all traded India symbols and normalize to yfinance NSE format."""
#     try:
#         equity_df = pd.read_csv(csv_path)
#     except Exception as e:
#         print(f"Warning: Could not read {csv_path}: {e}")
#         return []

#     symbol_col = next((c for c in equity_df.columns if c.strip().upper() == "SYMBOL"), None)
#     if symbol_col is None:
#         print(f"Warning: 'SYMBOL' column not found in {csv_path}")
#         return []

#     symbols = []
#     seen = set()
#     for raw in equity_df[symbol_col].dropna().astype(str).str.strip():
#         if not raw:
#             continue
#         symbol = raw if raw.endswith((".NS", ".BO")) else f"{raw}.NS"
#         if symbol not in seen:
#             symbols.append(symbol)
#             seen.add(symbol)

#     return symbols


# DEFAULT_SYMBOLS = load_default_symbols()

# ─────────────────────────────────────────────────────────────────────
# SCORING RULES — all thresholds in one place for easy tuning
# ─────────────────────────────────────────────────────────────────────

@dataclass
class RuleResult:
    rule_id: str
    rule_name: str
    source: str
    value: Optional[float]
    score: float
    max_score: float
    grade: str  # "excellent", "good", "fair", "poor"


def score_tiered(value, tiers, inverse=False):
    """
    Score a value against tiered thresholds.
    tiers: list of (threshold, points, grade) in descending priority.
    inverse: if True, lower values score higher.
    """
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return 0, "N/A"
    
    if inverse:
        for threshold, points, grade in tiers:
            if value <= threshold:
                return points, grade
        return 0, "poor"
    else:
        for threshold, points, grade in tiers:
            if value >= threshold:
                return points, grade
        return 0, "poor"


# ─────────────────────────────────────────────────────────────────────
# DATA FETCHING
# ─────────────────────────────────────────────────────────────────────

def safe_get(d, key, default=None):
    """Safely get a value from dict, handling None and NaN."""
    val = d.get(key, default)
    if val is None:
        return default
    if isinstance(val, float) and math.isnan(val):
        return default
    return val


def is_none_or_nan(value):
    """Return True for None/NaN values."""
    return value is None or (isinstance(value, float) and math.isnan(value))


def getMoneycontrolRatiosUrl(stockHtml: str) -> Optional[str]:
    """Extract Moneycontrol ratios page URL from stock quote HTML."""
    match = re.search(
        r'https://www\.moneycontrol\.com/financials/[^"\']+/ratiosVI/[^"\'#]+'
        , stockHtml,
    )
    return match.group(0) if match else None


def extractMoneycontrolNumeric(ratiosHtml: str, metricLabel: str) -> Optional[float]:
    """Extract latest numeric value from a named metric row in ratios HTML."""
    rowPattern = re.compile(
        r"<tr[^>]*>.*?<td[^>]*>\s*" + re.escape(metricLabel) + r"\s*</td>(.*?)</tr>",
        re.I | re.S,
    )
    rowMatch = rowPattern.search(ratiosHtml)
    if not rowMatch:
        return None

    valueCells = re.findall(r"<td[^>]*>\s*([^<]+?)\s*</td>", rowMatch.group(1), re.I | re.S)
    cleanValues = []
    for valueCell in valueCells:
        valueText = (
            valueCell.strip()
            .replace("&nbsp;", "")
            .replace(",", "")
        )
        if valueText and valueText not in ("-", "--", "NA", "N/A"):
            cleanValues.append(valueText)

    if not cleanValues:
        return None

    for valueText in cleanValues:
        try:
            return float(valueText)
        except Exception:
            continue
    return None


def throttleScreenerRequest() -> None:
    """Enforce a minimum 1-second gap between Screener.in requests."""
    global screenerLastRequestTs
    nowTs = time.time()
    elapsed = nowTs - screenerLastRequestTs
    if elapsed < SCREENER_MIN_INTERVAL_SECONDS:
        time.sleep(SCREENER_MIN_INTERVAL_SECONDS - elapsed)
    screenerLastRequestTs = time.time()


def parseScreenerNumeric(valueText: str) -> Optional[float]:
    """Parse Screener.in ratio value text into float."""
    cleaned = (
        valueText.replace("₹", "")
        .replace("%", "")
        .replace("Cr.", "")
        .replace(",", "")
        .strip()
    )
    match = re.search(r"-?\d+(?:\.\d+)?", cleaned)
    if not match:
        return None
    try:
        return float(match.group(0))
    except Exception:
        return None


def normalizeFallbackValue(fieldName: str, numericValue: float) -> float:
    """Normalize fallback values to yfinance-like scales when needed."""
    if fieldName == "returnOnEquity" and numericValue > 1:
        return numericValue / 100.0
    return numericValue


def getStatementMetric(frame: pd.DataFrame, candidateRows: list[str]) -> Optional[float]:
    """Return latest numeric metric from a statement frame by candidate row names."""
    if frame is None or frame.empty:
        return None

    rowLookup = {str(idx).strip().lower(): idx for idx in frame.index}
    selectedIndex = None
    for candidate in candidateRows:
        candidateKey = candidate.strip().lower()
        if candidateKey in rowLookup:
            selectedIndex = rowLookup[candidateKey]
            break
    if selectedIndex is None:
        return None

    rowSeries = frame.loc[selectedIndex]
    if isinstance(rowSeries, pd.Series):
        for value in rowSeries.tolist():
            if pd.notna(value):
                try:
                    return float(value)
                except Exception:
                    continue
    return None


def getStatementSeries(frame: pd.DataFrame, candidateRows: list[str]) -> list[float]:
    """Return numeric series for a statement row across periods (latest to oldest)."""
    if frame is None or frame.empty:
        return []

    rowLookup = {str(idx).strip().lower(): idx for idx in frame.index}
    selectedIndex = None
    for candidate in candidateRows:
        candidateKey = candidate.strip().lower()
        if candidateKey in rowLookup:
            selectedIndex = rowLookup[candidateKey]
            break
    if selectedIndex is None:
        return []

    rowSeries = frame.loc[selectedIndex]
    if not isinstance(rowSeries, pd.Series):
        return []

    values = []
    for value in rowSeries.tolist():
        if pd.notna(value):
            try:
                values.append(float(value))
            except Exception:
                continue
    return values


def computeCagrFromSeries(values: list[float]) -> Optional[float]:
    """Compute CAGR from values ordered latest to oldest; returns decimal growth."""
    if len(values) < 2:
        return None

    latestValue = values[0]
    oldestValue = values[-1]
    periods = len(values) - 1
    if periods <= 0 or latestValue <= 0 or oldestValue <= 0:
        return None

    try:
        return (latestValue / oldestValue) ** (1.0 / periods) - 1.0
    except Exception:
        return None


def fillDerivedYfinanceMetrics(ticker: yf.Ticker, rawInputs: dict, info: dict) -> None:
    """Fill missing key metrics from yfinance financial statements before web fallbacks."""
    try:
        incomeStatement = ticker.financials
    except Exception:
        incomeStatement = pd.DataFrame()
    try:
        balanceSheet = ticker.balance_sheet
    except Exception:
        balanceSheet = pd.DataFrame()
    try:
        cashFlow = ticker.cashflow
    except Exception:
        cashFlow = pd.DataFrame()

    marketCap = rawInputs.get("marketCap") if not is_none_or_nan(rawInputs.get("marketCap")) else None
    totalDebt = rawInputs.get("totalDebt") if not is_none_or_nan(rawInputs.get("totalDebt")) else None
    totalCash = rawInputs.get("totalCash") if not is_none_or_nan(rawInputs.get("totalCash")) else None
    ebitda = rawInputs.get("ebitda") if not is_none_or_nan(rawInputs.get("ebitda")) else None
    enterpriseValue = info.get("enterpriseValue") if not is_none_or_nan(info.get("enterpriseValue")) else None

    operatingIncome = getStatementMetric(incomeStatement, ["Operating Income", "Operating Income Or Loss"])
    grossProfit = getStatementMetric(incomeStatement, ["Gross Profit"])
    totalRevenue = getStatementMetric(incomeStatement, ["Total Revenue", "Operating Revenue", "Revenue"])
    netIncome = getStatementMetric(incomeStatement, ["Net Income", "Net Income Common Stockholders"])
    depreciation = getStatementMetric(
        cashFlow,
        [
            "Depreciation And Amortization",
            "Depreciation",
            "Depreciation Amortization Depletion",
        ],
    )

    if totalDebt is None:
        totalDebt = getStatementMetric(
            balanceSheet,
            [
                "Total Debt",
                "Total debt",
                "Long Term Debt",
                "Long Term Debt And Capital Lease Obligation",
                "Current Debt",
                "Current Debt And Capital Lease Obligation",
            ],
        )
    if totalCash is None:
        totalCash = getStatementMetric(balanceSheet, ["Cash And Cash Equivalents", "Cash", "Cash Cash Equivalents And Short Term Investments"])
    if ebitda is None:
        ebitda = getStatementMetric(incomeStatement, ["EBITDA", "Ebitda"])
    if ebitda is None and operatingIncome is not None and depreciation is not None:
        ebitda = float(operatingIncome) + abs(float(depreciation))
    if ebitda is None and operatingIncome is not None:
        # EBIT proxy when EBITDA is unavailable in yfinance statements.
        ebitda = float(operatingIncome)

    if is_none_or_nan(rawInputs.get("grossMargins")):
        if grossProfit is not None and totalRevenue and totalRevenue > 0:
            rawInputs["grossMargins"] = float(grossProfit) / float(totalRevenue)

    if is_none_or_nan(rawInputs.get("operatingMargins")):
        if operatingIncome is not None and totalRevenue and totalRevenue > 0:
            rawInputs["operatingMargins"] = float(operatingIncome) / float(totalRevenue)

    if is_none_or_nan(rawInputs.get("profitMargins")):
        if netIncome is not None and totalRevenue and totalRevenue > 0:
            rawInputs["profitMargins"] = float(netIncome) / float(totalRevenue)

    if is_none_or_nan(rawInputs.get("revenueGrowth")):
        revenueSeries = getStatementSeries(incomeStatement, ["Total Revenue", "Operating Revenue", "Revenue"])
        revenueCagr = computeCagrFromSeries(revenueSeries)
        if revenueCagr is not None:
            rawInputs["revenueGrowth"] = revenueCagr

    if is_none_or_nan(rawInputs.get("earningsGrowth")):
        earningsSeries = getStatementSeries(incomeStatement, ["Net Income", "Net Income Common Stockholders"])
        earningsCagr = computeCagrFromSeries(earningsSeries)
        if earningsCagr is not None:
            rawInputs["earningsGrowth"] = earningsCagr
        elif not is_none_or_nan(info.get("earningsQuarterlyGrowth")):
            rawInputs["earningsGrowth"] = float(info.get("earningsQuarterlyGrowth"))

    if is_none_or_nan(rawInputs.get("currentRatio")):
        currentAssets = getStatementMetric(balanceSheet, ["Current Assets", "Total Current Assets"])
        currentLiabilities = getStatementMetric(balanceSheet, ["Current Liabilities", "Total Current Liabilities"])
        if currentAssets and currentLiabilities and currentLiabilities > 0:
            rawInputs["currentRatio"] = float(currentAssets) / float(currentLiabilities)
        elif not is_none_or_nan(info.get("quickRatio")):
            rawInputs["currentRatio"] = float(info.get("quickRatio"))

    if is_none_or_nan(rawInputs.get("debtToEquity")):
        equityValue = getStatementMetric(balanceSheet, ["Stockholders Equity", "Total Stockholder Equity", "Common Stock Equity", "Total Equity Gross Minority Interest"])
        if totalDebt and equityValue and equityValue > 0:
            rawInputs["debtToEquity"] = float(totalDebt) / float(equityValue)
        else:
            totalLiabilities = getStatementMetric(balanceSheet, ["Total Liabilities Net Minority Interest", "Total Liabilities"])
            if totalLiabilities and equityValue and equityValue > 0:
                rawInputs["debtToEquity"] = float(totalLiabilities) / float(equityValue)

    if is_none_or_nan(rawInputs.get("returnOnEquity")):
        equityValue = getStatementMetric(balanceSheet, ["Stockholders Equity", "Total Stockholder Equity", "Common Stock Equity", "Total Equity Gross Minority Interest"])
        if netIncome is not None and equityValue and equityValue > 0:
            rawInputs["returnOnEquity"] = float(netIncome) / float(equityValue)

    if is_none_or_nan(rawInputs.get("pegRatio")):
        trailingPe = rawInputs.get("trailingPE")
        earningsGrowth = rawInputs.get("earningsGrowth")
        if (
            trailingPe is not None
            and not is_none_or_nan(trailingPe)
            and trailingPe > 0
            and earningsGrowth is not None
            and not is_none_or_nan(earningsGrowth)
        ):
            growthPercent = earningsGrowth * 100.0 if earningsGrowth <= 1 else earningsGrowth
            if growthPercent != 0:
                rawInputs["pegRatio"] = float(trailingPe) / float(growthPercent)

    if is_none_or_nan(rawInputs.get("enterpriseToEbitda")):
        if enterpriseValue is not None and ebitda and ebitda > 0:
            rawInputs["enterpriseToEbitda"] = float(enterpriseValue) / float(ebitda)
        elif marketCap and totalDebt is not None and totalCash is not None and ebitda and ebitda > 0:
            enterpriseValue = float(marketCap) + float(totalDebt) - float(totalCash)
            rawInputs["enterpriseToEbitda"] = enterpriseValue / float(ebitda)

    if is_none_or_nan(rawInputs.get("dividendYield")):
        currentPrice = rawInputs.get("currentPrice") if not is_none_or_nan(rawInputs.get("currentPrice")) else rawInputs.get("regularMarketPrice")
        annualDividendRate = info.get("trailingAnnualDividendRate")
        if is_none_or_nan(annualDividendRate):
            annualDividendRate = info.get("dividendRate")
        if (
            currentPrice is not None
            and not is_none_or_nan(currentPrice)
            and currentPrice > 0
            and annualDividendRate is not None
            and not is_none_or_nan(annualDividendRate)
        ):
            rawInputs["dividendYield"] = float(annualDividendRate) / float(currentPrice)

    if totalDebt is not None and is_none_or_nan(rawInputs.get("totalDebt")):
        rawInputs["totalDebt"] = totalDebt
    if totalCash is not None and is_none_or_nan(rawInputs.get("totalCash")):
        rawInputs["totalCash"] = totalCash
    if ebitda is not None and is_none_or_nan(rawInputs.get("ebitda")):
        rawInputs["ebitda"] = ebitda


def fetchScreenerTopRatios(symbol: str) -> tuple[dict[str, float], Optional[str], Optional[str]]:
    """Fetch numeric ratio-like metrics from Screener.in for NSE symbol."""
    if symbol in screenerTopRatiosCache:
        return screenerTopRatiosCache[symbol]

    nseSymbol = symbol.replace(".NS", "").replace(".BO", "")
    screenerUrl = f"https://www.screener.in/company/{nseSymbol}/consolidated/"
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
    }

    try:
        throttleScreenerRequest()
        response = screenerSession.get(
            screenerUrl,
            timeout=SCREENER_TIMEOUT_SECONDS,
            headers=headers,
        )
        if response.status_code != 200:
            result = ({}, screenerUrl, f"screener_http_{response.status_code}")
            screenerTopRatiosCache[symbol] = result
            return result

        topRatioPairs = re.findall(
            r'<li class="flex flex-space-between"[^>]*>\s*<span class="name">\s*(.*?)\s*</span>\s*<span class="nowrap value">\s*(.*?)\s*</span>\s*</li>',
            response.text,
            re.S,
        )
        topRatioValues = {}
        for nameHtml, valueHtml in topRatioPairs:
            metricName = " ".join(re.sub(r"<[^>]+>", " ", nameHtml).split()).strip().lower()
            metricValueText = " ".join(re.sub(r"<[^>]+>", " ", valueHtml).split()).strip()
            metricValue = parseScreenerNumeric(metricValueText)
            if metricValue is not None:
                topRatioValues[metricName] = metricValue

        # Parse any 2-column table row key/value pairs as extra candidates.
        rowPairs = re.findall(
            r"<tr[^>]*>\s*(?:<th[^>]*>|<td[^>]*>)(.*?)</(?:th|td)>\s*<td[^>]*>(.*?)</td>\s*</tr>",
            response.text,
            re.S,
        )
        for keyHtml, valueHtml in rowPairs:
            metricName = " ".join(re.sub(r"<[^>]+>", " ", keyHtml).split()).strip().lower()
            metricValueText = " ".join(re.sub(r"<[^>]+>", " ", valueHtml).split()).strip()
            metricValue = parseScreenerNumeric(metricValueText)
            if metricName and metricValue is not None and metricName not in topRatioValues:
                topRatioValues[metricName] = metricValue

        if not topRatioValues:
            result = ({}, screenerUrl, "screener_metrics_not_found")
            screenerTopRatiosCache[symbol] = result
            return result

        result = (topRatioValues, screenerUrl, None)
        screenerTopRatiosCache[symbol] = result
        return result
    except Exception as error:
        result = ({}, screenerUrl, f"screener_request_error:{type(error).__name__}")
        screenerTopRatiosCache[symbol] = result
        return result


def fetchScreenerFallbackValue(symbol: str, fieldName: str) -> tuple[Optional[float], Optional[str]]:
    """Get fallback value from Screener.in metrics for selected field."""
    topRatioValues, _, fallbackError = fetchScreenerTopRatios(symbol)
    fieldCandidates = {
        "returnOnEquity": [
            "roe",
            "return on equity / networth (%)",
            "return on networth / equity (%)",
        ],
        "currentRatio": [
            "current ratio",
            "current ratio (x)",
        ],
        "enterpriseToEbitda": [
            "ev / ebitda",
            "ev/ebitda",
            "ev multiple-ebitda",
            "ev multiple - ebitda",
        ],
        "debtToEquity": [
            "debt to equity",
            "debt to equity ratio",
            "debt-equity ratio",
            "total debt/equity (x)",
        ],
    }

    candidates = fieldCandidates.get(fieldName, [])
    for candidateKey in candidates:
        rawValue = topRatioValues.get(candidateKey)
        if rawValue is not None:
            return normalizeFallbackValue(fieldName, rawValue), None

    return None, (fallbackError or f"screener_{fieldName}_missing")


def fetchMoneycontrolRatiosHtml(symbol: str) -> tuple[Optional[str], Optional[str], Optional[str]]:
    """Fetch Moneycontrol ratios page HTML for symbol with caching."""
    if symbol in moneycontrolRatiosCache:
        return moneycontrolRatiosCache[symbol]

    querySymbol = symbol.replace(".NS", "").replace(".BO", "")
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        suggestResponse = moneycontrolSession.get(
            MONEYCONTROL_SUGGEST_URL,
            params={"classic": 1, "query": querySymbol, "type": 1, "format": "json"},
            timeout=MONEYCONTROL_TIMEOUT_SECONDS,
            headers=headers,
        )
        suggestItems = suggestResponse.json() if suggestResponse.status_code == 200 else []
        if not suggestItems:
            result = (None, None, "no_suggest_match")
            moneycontrolRatiosCache[symbol] = result
            return result

        stockUrl = suggestItems[0].get("link_src")
        if not stockUrl:
            result = (None, None, "missing_stock_url")
            moneycontrolRatiosCache[symbol] = result
            return result

        stockResponse = moneycontrolSession.get(
            stockUrl,
            timeout=MONEYCONTROL_TIMEOUT_SECONDS,
            headers=headers,
        )
        ratiosUrl = getMoneycontrolRatiosUrl(stockResponse.text)
        if not ratiosUrl:
            result = (None, stockUrl, "missing_ratios_url")
            moneycontrolRatiosCache[symbol] = result
            return result

        ratiosResponse = moneycontrolSession.get(
            ratiosUrl,
            timeout=MONEYCONTROL_TIMEOUT_SECONDS,
            headers=headers,
        )
        if ratiosResponse.status_code != 200:
            result = (None, ratiosUrl, f"ratios_http_{ratiosResponse.status_code}")
            moneycontrolRatiosCache[symbol] = result
            return result

        result = (ratiosResponse.text, ratiosUrl, None)
        moneycontrolRatiosCache[symbol] = result
        return result
    except Exception as error:
        result = (None, None, f"request_error:{type(error).__name__}")
        moneycontrolRatiosCache[symbol] = result
        return result


def fillMissingFromMoneycontrol(symbol: str, rawInputs: dict) -> dict:
    """Fill selected NaN yfinance fields from Moneycontrol ratios page."""
    fallbackFields = [
        "returnOnEquity",
        "currentRatio",
        "enterpriseToEbitda",
        "debtToEquity",
    ]
    requestedFields = [fieldName for fieldName in fallbackFields if is_none_or_nan(rawInputs.get(fieldName))]
    fieldSources = {
        fieldName: ("yfinance" if not is_none_or_nan(rawInputs.get(fieldName)) else "missing")
        for fieldName in fallbackFields
    }

    fallbackResult = {
        "moneycontrolAttempted": bool(requestedFields),
        "moneycontrolFallbackFailed": False,
        "moneycontrolFallbackError": "",
        "moneycontrolFailedFields": "",
        "moneycontrolFilledFields": "",
        "moneycontrolFilledCount": 0,
        "moneycontrolRequestedCount": len(requestedFields),
        "moneycontrolRatiosUrl": "",
        "sourceReturnOnEquity": fieldSources["returnOnEquity"],
        "sourceCurrentRatio": fieldSources["currentRatio"],
        "sourceEnterpriseToEbitda": fieldSources["enterpriseToEbitda"],
        "sourceDebtToEquity": fieldSources["debtToEquity"],
        "screenerFallbackFailed": False,
        "screenerFailedFields": "",
        "screenerFallbackError": "",
    }

    if not requestedFields:
        return fallbackResult

    ratiosHtml, ratiosUrl, fallbackError = fetchMoneycontrolRatiosHtml(symbol)
    fallbackResult["moneycontrolRatiosUrl"] = ratiosUrl or ""
    fallbackResult["moneycontrolFallbackError"] = fallbackError or ""

    filledFields = []
    failedFields = []
    screenerFailedFields = []
    screenerErrors = []
    fallbackErrors = [fallbackError] if fallbackError else []
    for fieldName in requestedFields:
        extractedValue = None
        fieldSource = ""
        if ratiosHtml:
            for metricLabel in moneycontrolFieldLabels.get(fieldName, []):
                extractedValue = extractMoneycontrolNumeric(ratiosHtml, metricLabel)
                if extractedValue is not None:
                    extractedValue = normalizeFallbackValue(fieldName, extractedValue)
                    fieldSource = "moneycontrol"
                    break

        if extractedValue is None and fieldName == "returnOnEquity":
            extractedValue, screenerError = fetchScreenerFallbackValue(symbol, fieldName)
            if extractedValue is not None:
                fieldSource = "screener"
            elif screenerError:
                screenerFailedFields.append(fieldName)
                screenerErrors.append(screenerError)
                fallbackErrors.append(screenerError)

        if extractedValue is None:
            failedFields.append(fieldName)
        else:
            rawInputs[fieldName] = extractedValue
            filledFields.append(fieldName)
            fieldSources[fieldName] = fieldSource or "moneycontrol"

    fallbackResult["moneycontrolFallbackFailed"] = len(failedFields) > 0
    fallbackResult["moneycontrolFallbackError"] = "|".join([e for e in fallbackErrors if e])
    fallbackResult["moneycontrolFailedFields"] = "|".join(failedFields)
    fallbackResult["moneycontrolFilledFields"] = "|".join(filledFields)
    fallbackResult["moneycontrolFilledCount"] = len(filledFields)
    fallbackResult["sourceReturnOnEquity"] = fieldSources["returnOnEquity"]
    fallbackResult["sourceCurrentRatio"] = fieldSources["currentRatio"]
    fallbackResult["sourceEnterpriseToEbitda"] = fieldSources["enterpriseToEbitda"]
    fallbackResult["sourceDebtToEquity"] = fieldSources["debtToEquity"]
    fallbackResult["screenerFallbackFailed"] = len(screenerFailedFields) > 0
    fallbackResult["screenerFailedFields"] = "|".join(screenerFailedFields)
    fallbackResult["screenerFallbackError"] = "|".join([e for e in screenerErrors if e])
    return fallbackResult


def fetch_stock_data(symbol: str) -> dict:
    """
    Fetch all fundamental + technical data for a single stock.
    Returns a flat dict with all fields needed for scoring.
    """
    try:
        ticker = yf.Ticker(symbol)
        info = ticker.info or {}

        raw_input_keys = [
            "marketCap", "grossMargins", "operatingMargins", "profitMargins", "returnOnEquity",
            "trailingPE", "pegRatio", "priceToBook", "enterpriseToEbitda", "debtToEquity",
            "currentRatio", "dividendYield", "revenueGrowth", "earningsGrowth", "trailingEps",
            "bookValue", "totalDebt", "totalCash", "ebitda", "currentPrice", "regularMarketPrice",
        ]
        raw_inputs = {
            key: (math.nan if is_none_or_nan(info.get(key)) else info.get(key))
            for key in raw_input_keys
        }

        # Fill what we can from yfinance statements first to reduce external fallback calls.
        fillDerivedYfinanceMetrics(ticker, raw_inputs, info)
        if USE_EXTERNAL_FALLBACK:
            moneycontrolFallback = fillMissingFromMoneycontrol(symbol, raw_inputs)
        else:
            moneycontrolFallback = {
                "moneycontrolAttempted": False,
                "moneycontrolFallbackFailed": False,
                "moneycontrolFallbackError": "",
                "moneycontrolFailedFields": "",
                "moneycontrolFilledFields": "",
                "moneycontrolFilledCount": 0,
                "moneycontrolRequestedCount": 0,
                "moneycontrolRatiosUrl": "",
                "sourceReturnOnEquity": "yfinance",
                "sourceCurrentRatio": "yfinance",
                "sourceEnterpriseToEbitda": "yfinance",
                "sourceDebtToEquity": "yfinance",
                "screenerFallbackFailed": False,
                "screenerFailedFields": "",
                "screenerFallbackError": "",
            }

        def getRawNumeric(key: str, default: float = 0):
            value = raw_inputs.get(key, default)
            return default if is_none_or_nan(value) else value
        
        if not info.get("regularMarketPrice") and not info.get("currentPrice"):
            return {"symbol": symbol, "error": "No data available"}
        
        # ── Price & Technical Data ──
        hist = ticker.history(period="1y")
        technicals = compute_technicals(hist)
        technicals = {k: (math.nan if is_none_or_nan(v) else v) for k, v in technicals.items()}
        
        # ── Fundamentals from info ──
        market_cap = getRawNumeric("marketCap", 0)
        
        gross_margins = getRawNumeric("grossMargins", 0)
        operating_margins = getRawNumeric("operatingMargins", 0)
        profit_margins = getRawNumeric("profitMargins", 0)
        roe = getRawNumeric("returnOnEquity", 0)
        
        trailing_pe = getRawNumeric("trailingPE", 0)
        peg = getRawNumeric("pegRatio", 0)
        pb = getRawNumeric("priceToBook", 0)
        ev_ebitda = getRawNumeric("enterpriseToEbitda", 0)
        de_ratio = getRawNumeric("debtToEquity", 0)
        current_ratio = getRawNumeric("currentRatio", 0)
        div_yield = getRawNumeric("dividendYield", 0)
        revenue_growth = getRawNumeric("revenueGrowth", 0)
        earnings_growth = getRawNumeric("earningsGrowth", 0)
        
        trailing_eps = getRawNumeric("trailingEps", 0)
        book_value = getRawNumeric("bookValue", 0)
        
        # ── Derived Calculations ──
        gross_margin = gross_margins * 100 if gross_margins else 0
        operating_margin = operating_margins * 100 if operating_margins else 0
        net_margin = profit_margins * 100 if profit_margins else 0
        roe_pct = roe * 100 if roe else 0
        de = de_ratio / 100 if de_ratio and de_ratio > 5 else de_ratio  # yfinance sometimes returns as %
        dividend_yield_pct = div_yield * 100 if div_yield else 0
        revenue_growth_pct = revenue_growth * 100 if revenue_growth else 0
        earnings_growth_pct = earnings_growth * 100 if earnings_growth else 0
        
        # Earnings Yield
        earnings_yield = (1 / trailing_pe * 100) if trailing_pe and trailing_pe > 0 else 0
        
        # Graham Number
        graham_number = 0
        graham_ratio = 999
        if trailing_eps and trailing_eps > 0 and book_value and book_value > 0:
            graham_number = math.sqrt(22.5 * trailing_eps * book_value)
            current_price = getRawNumeric("currentPrice", 0) or getRawNumeric("regularMarketPrice", 0)
            if current_price and graham_number > 0:
                graham_ratio = current_price / graham_number
        
        # FCF data from cash flow statements
        fcf_data = compute_fcf_metrics(ticker)
        
        # Interest coverage
        interest_coverage = compute_interest_coverage(ticker)
        
        # Net Debt / EBITDA
        total_debt = getRawNumeric("totalDebt", 0) or 0
        total_cash = getRawNumeric("totalCash", 0) or 0
        ebitda = getRawNumeric("ebitda", 0) or 0
        net_debt = total_debt - total_cash
        net_debt_ebitda = net_debt / ebitda if ebitda and ebitda > 0 else 999
        
        # Profitable years estimate
        profitable_years = estimate_profitable_years(ticker)
        
        data = {
            "symbol": symbol,
            "name": safe_get(info, "longName", symbol),
            "sector": safe_get(info, "sector", ""),
            "industry": safe_get(info, "industry", ""),
            "market_cap_cr": market_cap / 1e7 if market_cap else 0,
            "current_price": getRawNumeric("currentPrice", 0) or getRawNumeric("regularMarketPrice", 0),
            
            # Profitability
            "gross_margin": gross_margin,
            "operating_margin": operating_margin,
            "net_margin": net_margin,
            "roe": roe_pct,
            "roce": roe_pct * 0.85,  # Approximation; real ROCE needs EBIT/(TA-CL)
            "eps_growth_5yr": earnings_growth_pct,  # Approximation
            "revenue_growth_3yr": revenue_growth_pct,
            "fcf_margin": fcf_data.get("fcf_margin", 0),
            
            # Balance Sheet
            "debt_to_equity": de if de else 0,
            "current_ratio": current_ratio,
            "interest_coverage": interest_coverage,
            "net_debt_to_ebitda": net_debt_ebitda,
            "fcf_positive_years": fcf_data.get("fcf_positive_years", 0),
            
            # Valuation
            "pe_ratio": trailing_pe,
            "peg_ratio": peg,
            "pb_ratio": pb,
            "ev_to_ebitda": ev_ebitda,
            "earnings_yield": earnings_yield,
            "dividend_yield": dividend_yield_pct,
            "graham_number": graham_number,
            "graham_number_ratio": graham_ratio,
            
            # Quality
            "profitable_years": profitable_years,
            "dividend_years": 5 if dividend_yield_pct > 0 else 0,  # Rough estimate
            "promoter_holding": None,  # Needs manual/CSV input
            "promoter_pledge": None,
            
            # Error
            "error": None,

            # Source completeness helper
            "histRows": int(len(hist)) if hist is not None else 0,

            # Moneycontrol fallback diagnostics
            "moneycontrolAttempted": moneycontrolFallback["moneycontrolAttempted"],
            "moneycontrolFallbackFailed": moneycontrolFallback["moneycontrolFallbackFailed"],
            "moneycontrolFallbackError": moneycontrolFallback["moneycontrolFallbackError"],
            "moneycontrolFailedFields": moneycontrolFallback["moneycontrolFailedFields"],
            "moneycontrolFilledFields": moneycontrolFallback["moneycontrolFilledFields"],
            "moneycontrolFilledCount": moneycontrolFallback["moneycontrolFilledCount"],
            "moneycontrolRequestedCount": moneycontrolFallback["moneycontrolRequestedCount"],
            "moneycontrolRatiosUrl": moneycontrolFallback["moneycontrolRatiosUrl"],
            "sourceReturnOnEquity": moneycontrolFallback["sourceReturnOnEquity"],
            "sourceCurrentRatio": moneycontrolFallback["sourceCurrentRatio"],
            "sourceEnterpriseToEbitda": moneycontrolFallback["sourceEnterpriseToEbitda"],
            "sourceDebtToEquity": moneycontrolFallback["sourceDebtToEquity"],
            "screenerFallbackFailed": moneycontrolFallback["screenerFallbackFailed"],
            "screenerFailedFields": moneycontrolFallback["screenerFailedFields"],
            "screenerFallbackError": moneycontrolFallback["screenerFallbackError"],
        }
        data.update(raw_inputs)
        data.update(technicals)
        return data
        
    except Exception as e:
        return {"symbol": symbol, "error": str(e)}


def load_data_cache(cache_file: str) -> dict[str, dict]:
    """Load per-symbol fetched data cache from Excel."""
    cache_path = Path(cache_file)
    if not cache_path.exists():
        return {}
    try:
        cache_df = pd.read_excel(cache_path)
    except Exception as e:
        print(f"Warning: Could not read cache file {cache_file}: {e}")
        return {}

    if "symbol" not in cache_df.columns:
        return {}

    cache = {}
    for _, row in cache_df.iterrows():
        row_dict = {k: (None if pd.isna(v) else v) for k, v in row.to_dict().items()}
        symbol = row_dict.get("symbol")
        if symbol:
            cache[str(symbol)] = row_dict
    return cache


def save_data_cache(cache: dict[str, dict], cache_file: str):
    """Persist per-symbol fetched data cache to Excel."""
    if not cache:
        return
    cache_df = pd.DataFrame(cache.values())
    cache_df = cache_df.sort_values("symbol").reset_index(drop=True)
    cache_df.to_excel(cache_file, index=False)


def is_cache_fresh(cache_row: dict, cache_max_age_hours: float) -> bool:
    """Return True when cache row is recent enough to reuse."""
    ts = cache_row.get("_cache_timestamp")
    if not ts:
        return False
    try:
        ts_dt = datetime.fromisoformat(str(ts))
    except Exception:
        return False

    age_hours = (datetime.now() - ts_dt).total_seconds() / 3600
    return age_hours <= cache_max_age_hours


def strip_cache_meta(data: dict) -> dict:
    """Remove internal cache metadata keys before scoring."""
    return {k: v for k, v in data.items() if not str(k).startswith("_")}


def compute_technicals(hist: pd.DataFrame) -> dict:
    """Compute all technical indicators from price history."""
    if hist.empty or len(hist) < 50:
        return {
            "above_200sma": None, "above_50sma": None, "golden_alignment": None,
            "rsi_14": None, "volume_ratio": None, "pct_from_52w_high": None,
            "macd_bullish": None,
        }

    close = hist["Close"].dropna()
    volume = hist["Volume"].dropna()
    high = hist["High"].dropna() if "High" in hist.columns else pd.Series(dtype=float)
    if close.empty:
        return {
            "above_200sma": None, "above_50sma": None, "golden_alignment": None,
            "rsi_14": None, "volume_ratio": None, "pct_from_52w_high": None,
            "macd_bullish": None,
        }

    sma_200 = close.rolling(200).mean().iloc[-1] if len(close) >= 200 else close.mean()
    sma_50 = close.rolling(50).mean().iloc[-1] if len(close) >= 50 else close.mean()
    ema_21 = close.ewm(span=21).mean().iloc[-1]
    price = close.iloc[-1]
    
    # RSI
    delta = close.diff()
    gain = delta.where(delta > 0, 0).rolling(14).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(14).mean()
    rs = gain / loss
    rsi = (100 - (100 / (1 + rs))).iloc[-1]
    
    # Volume
    if volume.empty:
        vol_ratio = 1
    else:
        vol_avg = volume.rolling(20).mean().iloc[-1] if len(volume) >= 20 else volume.mean()
        vol_ratio = volume.iloc[-1] / vol_avg if vol_avg and vol_avg > 0 else 1
    
    # 52-week high
    high_52w = high.max() if not high.empty else close.max()
    pct_from_high = ((high_52w - price) / high_52w) * 100 if high_52w and high_52w > 0 else 0
    
    # MACD
    ema_12 = close.ewm(span=12).mean()
    ema_26 = close.ewm(span=26).mean()
    macd = ema_12 - ema_26
    signal = macd.ewm(span=9).mean()
    macd_bull = macd.iloc[-1] > signal.iloc[-1]
    
    return {
        "above_200sma": price > sma_200,
        "above_50sma": price > sma_50,
        "golden_alignment": (ema_21 > sma_50) and (sma_50 > sma_200) if len(close) >= 200 else False,
        "rsi_14": rsi if not math.isnan(rsi) else 50,
        "volume_ratio": vol_ratio,
        "pct_from_52w_high": pct_from_high,
        "macd_bullish": macd_bull,
    }


def compute_fcf_metrics(ticker) -> dict:
    """Extract FCF margin and positive-year count from cash flow statements."""
    try:
        cf = ticker.cashflow
        fin = ticker.financials
        if cf is None or cf.empty:
            return {"fcf_margin": 0, "fcf_positive_years": 0}
        
        # Free Cash Flow = Operating CF - CapEx
        ocf = cf.loc["Operating Cash Flow"] if "Operating Cash Flow" in cf.index else None
        capex = cf.loc["Capital Expenditure"] if "Capital Expenditure" in cf.index else None
        
        if ocf is None:
            return {"fcf_margin": 0, "fcf_positive_years": 0}
        
        if capex is not None:
            fcf = ocf + capex  # CapEx is typically negative
        else:
            fcf = ocf
        
        positive_years = (fcf > 0).sum()
        
        # FCF Margin (latest year)
        revenue = None
        if fin is not None and not fin.empty and "Total Revenue" in fin.index:
            revenue = fin.loc["Total Revenue"].iloc[0]
        
        fcf_margin = 0
        if revenue and revenue > 0 and len(fcf) > 0:
            fcf_margin = (fcf.iloc[0] / revenue) * 100
        
        return {"fcf_margin": fcf_margin, "fcf_positive_years": int(positive_years)}
    except:
        return {"fcf_margin": 0, "fcf_positive_years": 0}


def compute_interest_coverage(ticker) -> float:
    """EBIT / Interest Expense."""
    try:
        fin = ticker.financials
        if fin is None or fin.empty:
            return 99  # Assume no debt
        ebit = fin.loc["EBIT"].iloc[0] if "EBIT" in fin.index else None
        interest = fin.loc["Interest Expense"].iloc[0] if "Interest Expense" in fin.index else None
        if ebit and interest and interest < 0:
            return abs(ebit / interest)
        return 99  # No interest expense = great
    except:
        return 99


def estimate_profitable_years(ticker) -> int:
    """Count how many of the last available annual periods had positive net income."""
    try:
        fin = ticker.financials
        if fin is None or fin.empty:
            return 0
        if "Net Income" in fin.index:
            ni = fin.loc["Net Income"]
            return int((ni > 0).sum())
        return 0
    except:
        return 0


# ─────────────────────────────────────────────────────────────────────
# SCORING ENGINE
# ─────────────────────────────────────────────────────────────────────

def score_profitability(d: dict) -> tuple[float, list[RuleResult]]:
    """Score profitability rules (max 30 points)."""
    results = []
    total = 0
    
    rules = [
        ("P1", "Gross Margin", "Buffett", d.get("gross_margin", 0), 5,
         [(60, 5, "excellent"), (40, 3, "good"), (20, 1, "fair")], False),
        ("P2", "Net Margin", "Buffett", d.get("net_margin", 0), 4,
         [(25, 4, "excellent"), (20, 3, "good"), (10, 1, "fair")], False),
        ("P3", "Operating Margin", "Buffett", d.get("operating_margin", 0), 4,
         [(25, 4, "excellent"), (15, 3, "good"), (10, 1, "fair")], False),
        ("P4", "ROE", "Buffett", d.get("roe", 0), 5,
         [(25, 5, "excellent"), (20, 3, "good"), (15, 1, "fair")], False),
        ("P5", "ROCE", "Buffett/Graham", d.get("roce", 0), 4,
         [(25, 4, "excellent"), (15, 3, "good"), (10, 1, "fair")], False),
        ("P6", "EPS Growth", "Lynch", d.get("eps_growth_5yr", 0), 4,
         [(20, 4, "excellent"), (15, 3, "good"), (10, 1, "fair")], False),
        ("P7", "Revenue Growth 3Y", "Lynch", d.get("revenue_growth_3yr", 0), 2,
         [(20, 2, "excellent"), (10, 1.5, "good"), (5, 0.5, "fair")], False),
        ("P8", "FCF Margin", "Buffett", d.get("fcf_margin", 0), 2,
         [(15, 2, "excellent"), (8, 1.5, "good"), (3, 0.5, "fair")], False),
    ]
    
    for rid, name, source, value, max_pts, tiers, inv in rules:
        pts, grade = score_tiered(value, tiers, inverse=inv)
        # Cap eps growth > 30% at 2 points (unsustainable per Lynch)
        if rid == "P6" and value and value > 30:
            pts = min(pts, 2)
            grade = "caution"
        total += pts
        results.append(RuleResult(rid, name, source, value, pts, max_pts, grade))
    
    return total, results


def score_balance_sheet(d: dict) -> tuple[float, list[RuleResult]]:
    """Score balance sheet rules (max 20 points)."""
    results = []
    total = 0
    
    # B1: Debt-to-Equity (inverse)
    de = d.get("debt_to_equity", 999)
    if de is None: de = 999
    pts, grade = score_tiered(de, [(0.1, 5, "excellent"), (0.5, 4, "good"), (1.0, 2, "fair")], inverse=True)
    total += pts
    results.append(RuleResult("B1", "Debt/Equity", "Buffett/Graham", de, pts, 5, grade))
    
    # B2: Current Ratio
    cr = d.get("current_ratio", 0)
    pts, grade = score_tiered(cr, [(2.5, 4, "excellent"), (2.0, 3, "good"), (1.5, 2, "fair")], False)
    total += pts
    results.append(RuleResult("B2", "Current Ratio", "Graham", cr, pts, 4, grade))
    
    # B3: Interest Coverage
    ic = d.get("interest_coverage", 0)
    pts, grade = score_tiered(ic, [(10, 4, "excellent"), (5, 3, "good"), (3, 1, "fair")], False)
    total += pts
    results.append(RuleResult("B3", "Interest Coverage", "Graham", ic, pts, 4, grade))
    
    # B4: Net Debt/EBITDA (inverse)
    nd = d.get("net_debt_to_ebitda", 999)
    if nd is None: nd = 999
    pts, grade = score_tiered(nd, [(0, 4, "excellent"), (1.0, 3, "good"), (2.0, 1, "fair")], inverse=True)
    total += pts
    results.append(RuleResult("B4", "Net Debt/EBITDA", "Lynch", nd, pts, 4, grade))
    
    # B5: FCF Positive Years
    fy = d.get("fcf_positive_years", 0)
    pts, grade = score_tiered(fy, [(5, 3, "excellent"), (4, 2, "good"), (3, 1, "fair")], False)
    total += pts
    results.append(RuleResult("B5", "FCF Positive Years", "Buffett", fy, pts, 3, grade))
    
    return total, results


def score_valuation(d: dict) -> tuple[float, list[RuleResult]]:
    """Score valuation rules (max 25 points)."""
    results = []
    total = 0
    
    # V1: P/E (inverse)
    pe = d.get("pe_ratio", 0)
    if pe and pe > 0:
        pts, grade = score_tiered(pe, [(12, 5, "excellent"), (20, 3, "good"), (30, 1, "fair")], inverse=True)
    else:
        pts, grade = 0, "loss-making"
    total += pts
    results.append(RuleResult("V1", "P/E Ratio", "Graham", pe, pts, 5, grade))
    
    # V2: PEG (inverse)
    peg = d.get("peg_ratio", 0)
    if peg and peg > 0:
        pts, grade = score_tiered(peg, [(0.5, 5, "excellent"), (1.0, 4, "good"), (1.5, 2, "fair")], inverse=True)
    else:
        pts, grade = 0, "N/A"
    total += pts
    results.append(RuleResult("V2", "PEG Ratio", "Lynch", peg, pts, 5, grade))
    
    # V3: P/B (inverse)
    pb = d.get("pb_ratio", 0)
    if pb and pb > 0:
        pts, grade = score_tiered(pb, [(1.5, 4, "excellent"), (3.0, 2, "good"), (5.0, 1, "fair")], inverse=True)
    else:
        pts, grade = 0, "N/A"
    total += pts
    results.append(RuleResult("V3", "P/B Ratio", "Graham", pb, pts, 4, grade))
    
    # V4: Graham Number Ratio (inverse)
    gr = d.get("graham_number_ratio", 999)
    if gr and gr < 900:
        pts, grade = score_tiered(gr, [(0.8, 3, "excellent"), (1.0, 2, "good"), (1.2, 1, "fair")], inverse=True)
    else:
        pts, grade = 0, "N/A"
    total += pts
    results.append(RuleResult("V4", "Graham Number", "Graham", gr, pts, 3, grade))
    
    # V5: EV/EBITDA (inverse)
    ev = d.get("ev_to_ebitda", 0)
    if ev and ev > 0:
        pts, grade = score_tiered(ev, [(8, 3, "excellent"), (12, 2, "good"), (18, 1, "fair")], inverse=True)
    else:
        pts, grade = 0, "N/A"
    total += pts
    results.append(RuleResult("V5", "EV/EBITDA", "Buffett", ev, pts, 3, grade))
    
    # V6: Earnings Yield
    ey = d.get("earnings_yield", 0)
    pts, grade = score_tiered(ey, [(10, 3, "excellent"), (7, 2, "good"), (4, 1, "fair")], False)
    total += pts
    results.append(RuleResult("V6", "Earnings Yield", "Graham", ey, pts, 3, grade))
    
    # V7: Dividend Yield
    dy = d.get("dividend_yield", 0)
    pts, grade = score_tiered(dy, [(3.0, 2, "excellent"), (1.5, 1.5, "good"), (0.5, 0.5, "fair")], False)
    total += pts
    results.append(RuleResult("V7", "Dividend Yield", "Graham", dy, pts, 2, grade))
    
    return total, results


def score_quality(d: dict) -> tuple[float, list[RuleResult]]:
    """Score consistency/quality rules (max 15 points)."""
    results = []
    total = 0
    
    # Q1: Profitable years
    py = d.get("profitable_years", 0)
    pts, grade = score_tiered(py, [(4, 4, "excellent"), (3, 3, "good"), (2, 1, "fair")], False)
    total += pts
    results.append(RuleResult("Q1", "Profitable Years", "Graham", py, pts, 4, grade))
    
    # Q2: Dividend continuity
    dy = d.get("dividend_years", 0)
    pts, grade = score_tiered(dy, [(10, 3, "excellent"), (5, 2, "good"), (3, 1, "fair")], False)
    total += pts
    results.append(RuleResult("Q2", "Dividend History", "Graham", dy, pts, 3, grade))
    
    # Q3: Promoter holding
    ph = d.get("promoter_holding")
    if ph is not None:
        pts, grade = score_tiered(ph, [(65, 3, "excellent"), (50, 2, "good"), (35, 1, "fair")], False)
    else:
        pts, grade = 1.5, "unknown"  # Neutral if no data
    total += pts
    results.append(RuleResult("Q3", "Promoter Holding", "India", ph, pts, 3, grade))
    
    # Q4: Promoter pledge (inverse)
    pp = d.get("promoter_pledge")
    if pp is not None:
        if pp > 25:
            pts, grade = -2, "CRITICAL"
        elif pp > 10:
            pts, grade = 0, "poor"
        elif pp > 5:
            pts, grade = 1, "fair"
        elif pp > 1:
            pts, grade = 2, "good"
        else:
            pts, grade = 3, "excellent"
    else:
        pts, grade = 1.5, "unknown"
    total += pts
    results.append(RuleResult("Q4", "Promoter Pledge", "India", pp, pts, 3, grade))
    
    # Q5: Market Cap adequacy
    mc = d.get("market_cap_cr", 0)
    pts, grade = score_tiered(mc, [(10000, 2, "excellent"), (3000, 1.5, "good"), (1000, 1, "fair")], False)
    total += pts
    results.append(RuleResult("Q5", "Market Cap", "Graham", mc, pts, 2, grade))
    
    return total, results


def score_technicals(d: dict) -> tuple[float, list[RuleResult]]:
    """Score technical setup rules (max 10 points)."""
    results = []
    total = 0
    
    # T1: Price > 200 SMA
    val = d.get("above_200sma")
    pts = 2 if val else 0
    total += pts
    results.append(RuleResult("T1", "Above 200 SMA", "Technical", val, pts, 2, "yes" if val else "no"))
    
    # T2: Price > 50 SMA
    val = d.get("above_50sma")
    pts = 1.5 if val else 0
    total += pts
    results.append(RuleResult("T2", "Above 50 SMA", "Technical", val, pts, 1.5, "yes" if val else "no"))
    
    # T3: Golden alignment
    val = d.get("golden_alignment")
    pts = 2 if val else 0
    total += pts
    results.append(RuleResult("T3", "Golden Alignment", "Technical", val, pts, 2, "yes" if val else "no"))
    
    # T4: RSI in buy zone
    rsi = d.get("rsi_14", 50)
    if rsi is None: rsi = 50
    if 40 <= rsi <= 60:
        pts, grade = 2, "ideal"
    elif 30 <= rsi < 40 or 60 < rsi <= 70:
        pts, grade = 1, "acceptable"
    elif rsi < 30:
        pts, grade = 0.5, "oversold"
    else:
        pts, grade = 0, "overbought"
    total += pts
    results.append(RuleResult("T4", "RSI Zone", "Technical", rsi, pts, 2, grade))
    
    # T5: Volume surge
    vr = d.get("volume_ratio", 1)
    if vr is None: vr = 1
    if vr >= 2.0:
        pts, grade = 1.5, "strong"
    elif vr >= 1.5:
        pts, grade = 1, "moderate"
    elif vr >= 1.0:
        pts, grade = 0.5, "normal"
    else:
        pts, grade = 0, "weak"
    total += pts
    results.append(RuleResult("T5", "Volume Surge", "Technical", vr, pts, 1.5, grade))
    
    # T6: Near 52-week high
    pct = d.get("pct_from_52w_high", 100)
    if pct is None: pct = 100
    if pct <= 10:
        pts, grade = 1, "near_high"
    elif pct <= 20:
        pts, grade = 0.5, "moderate"
    else:
        pts, grade = 0, "far"
    total += pts
    results.append(RuleResult("T6", "Near 52W High", "Technical", pct, pts, 1, grade))
    
    return total, results


def check_red_flags(d: dict) -> list[str]:
    """Return list of red flag strings. Each flag carries a -10 penalty."""
    flags = []
    
    pp = d.get("promoter_pledge")
    if pp is not None and pp > 25:
        flags.append("🚨 CRITICAL_PLEDGE_RISK")
    
    de = d.get("debt_to_equity", 0)
    if de and de > 2.0:
        flags.append("🚨 EXCESSIVE_DEBT")
    
    pe = d.get("pe_ratio", 0)
    if pe and pe < 0:
        flags.append("🚨 LOSS_MAKING")
    
    cr = d.get("current_ratio", 0)
    if cr and cr < 1.0 and cr > 0:
        flags.append("🚨 LIQUIDITY_CRISIS")
    
    mc = d.get("market_cap_cr", 0)
    if mc and mc < 200:
        flags.append("⚠️ MICRO_CAP")
    
    return flags


def score_stock(data: dict) -> dict:
    """
    Master scoring function. Returns complete analysis.
    """
    if data.get("error"):
        return {
            "symbol": data["symbol"],
            "error": data["error"],
            "total_score": 0,
            "grade": "Error",
        }
    
    prof_score, prof_details = score_profitability(data)
    bs_score, bs_details = score_balance_sheet(data)
    val_score, val_details = score_valuation(data)
    qual_score, qual_details = score_quality(data)
    tech_score, tech_details = score_technicals(data)
    
    red_flags = check_red_flags(data)
    penalty = len(red_flags) * 10
    
    raw_total = prof_score + bs_score + val_score + qual_score + tech_score
    total_raw = max(0, raw_total - penalty)

    import_keys = [
        "marketCap", "grossMargins", "operatingMargins", "profitMargins", "returnOnEquity",
        "trailingPE", "pegRatio", "priceToBook", "enterpriseToEbitda", "debtToEquity",
        "currentRatio", "dividendYield", "revenueGrowth", "earningsGrowth", "trailingEps",
        "bookValue", "totalDebt", "totalCash", "ebitda", "currentPrice", "regularMarketPrice",
        "above_200sma", "above_50sma", "golden_alignment", "rsi_14", "volume_ratio", "pct_from_52w_high", "macd_bullish",
        "histRows",
    ]
    import_values = {key: data.get(key, math.nan) for key in import_keys}

    coverage_keys = [k for k in import_keys if k != "histRows"]
    data_coverage_count = sum(0 if is_none_or_nan(data.get(k, math.nan)) else 1 for k in coverage_keys)
    data_coverage_total = len(coverage_keys)
    data_coverage_pct = (data_coverage_count / data_coverage_total * 100.0) if data_coverage_total else 0.0
    score_out_of = round(100.0 * (data_coverage_pct / 100.0), 1)
    coverageAdjustedScore = round(total_raw * (data_coverage_pct / 100.0), 1)

    ruleDetails = prof_details + bs_details + val_details + qual_details + tech_details
    availableRules = [r for r in ruleDetails if not is_none_or_nan(r.value)]
    availableMaxScore = sum(r.max_score for r in availableRules)
    availableEarnedScore = sum(r.score for r in availableRules)
    intrinsicScore = (
        round(100.0 * availableEarnedScore / availableMaxScore, 1)
        if availableMaxScore > 0
        else 0.0
    )
    normalizedScore = intrinsicScore

    confidenceWeights = {
        "marketCap": 3.0,
        "grossMargins": 2.0,
        "operatingMargins": 2.0,
        "profitMargins": 2.0,
        "returnOnEquity": 3.0,
        "trailingPE": 3.0,
        "pegRatio": 2.0,
        "priceToBook": 2.0,
        "enterpriseToEbitda": 3.0,
        "debtToEquity": 3.0,
        "currentRatio": 3.0,
        "dividendYield": 1.5,
        "revenueGrowth": 2.0,
        "earningsGrowth": 2.0,
        "trailingEps": 1.5,
        "bookValue": 1.5,
        "totalDebt": 1.5,
        "totalCash": 1.0,
        "ebitda": 2.0,
        "currentPrice": 1.0,
        "regularMarketPrice": 1.0,
        "above_200sma": 1.0,
        "above_50sma": 1.0,
        "golden_alignment": 1.0,
        "rsi_14": 1.0,
        "volume_ratio": 1.0,
        "pct_from_52w_high": 1.0,
        "macd_bullish": 1.0,
    }
    weightTotal = sum(confidenceWeights.values())
    weightedAvailable = sum(
        weight for key, weight in confidenceWeights.items()
        if not is_none_or_nan(data.get(key, math.nan))
    )
    dataConfidence = round(100.0 * weightedAvailable / weightTotal, 1) if weightTotal > 0 else 0.0
    finalScore = round(intrinsicScore * (0.5 + 0.5 * (dataConfidence / 100.0)), 1)
    
    # Grade
    if normalizedScore >= 80:
        grade = "🟢 Strong Buy"
    elif normalizedScore >= 70:
        grade = "🟢 Buy"
    elif normalizedScore >= 60:
        grade = "🟡 Watchlist"
    elif normalizedScore >= 50:
        grade = "🟠 Hold"
    else:
        grade = "🔴 Avoid"
    
    return {
        "symbol": data.get("symbol", ""),
        "name": data.get("name", ""),
        "sector": data.get("sector", ""),
        "price": data.get("current_price", 0),
        "market_cap_cr": data.get("market_cap_cr", 0),
        "total_score": normalizedScore,
        "coverage_adjusted_score": coverageAdjustedScore,
        "total_score_raw": round(total_raw, 1),
        "score_out_of": score_out_of,
        "data_coverage_pct": round(data_coverage_pct, 1),
        "data_coverage_count": data_coverage_count,
        "data_coverage_total": data_coverage_total,
        "normalizedScore": normalizedScore,
        "intrinsicScore": intrinsicScore,
        "dataConfidence": dataConfidence,
        "finalScore": finalScore,
        "grade": grade,
        "profitability": round(prof_score, 1),
        "balance_sheet": round(bs_score, 1),
        "valuation": round(val_score, 1),
        "quality": round(qual_score, 1),
        "technicals": round(tech_score, 1),
        "red_flags": " | ".join(red_flags) if red_flags else "",
        "penalty": penalty,
        "pe": data.get("pe_ratio", 0),
        "roe": data.get("roe", 0),
        "de": data.get("debt_to_equity", 0),
        "rsi": data.get("rsi_14", 0),
        "details": ruleDetails,
        "error": None,
        **import_values,
    }


# ─────────────────────────────────────────────────────────────────────
# MAIN RUNNER
# ─────────────────────────────────────────────────────────────────────

def run_screener(
    symbols: list[str],
    verbose: bool = True,
    use_cache: bool = True,
    cache_mode: str = "auto",
    cache_file: str = "yfinance_cache.xlsx",
    cache_max_age_hours: float = 24,
    refresh_cache: bool = False,
) -> pd.DataFrame:
    """
    Run the full screener on a list of symbols.
    Returns a DataFrame sorted by total_score descending.
    """
    results = []
    total = len(symbols)
    cache = load_data_cache(cache_file) if use_cache else {}
    cache_dirty = False
    
    for i, symbol in enumerate(symbols):
        source = "fetch"
        if verbose:
            print(f"  [{i+1}/{total}] Fetching {symbol}...", end=" ", flush=True)

        cached_row = cache.get(symbol) if use_cache else None
        should_use_cache_row = False
        if cached_row and not refresh_cache:
            if cache_mode == "on":
                should_use_cache_row = True
            elif cache_mode == "auto" and is_cache_fresh(cached_row, cache_max_age_hours):
                should_use_cache_row = True

        if should_use_cache_row:
            data = strip_cache_meta(cached_row)
            source = "cache"
        else:
            data = fetch_stock_data(symbol)
            if use_cache and not data.get("error"):
                cache[symbol] = {
                    **data,
                    "_cache_timestamp": datetime.now().isoformat(timespec="seconds"),
                }
                cache_dirty = True

            # If fresh pull failed, fall back to any stale cache row.
            if data.get("error") and cached_row:
                data = strip_cache_meta(cached_row)
                source = "stale-cache"

        scored = score_stock(data)
        results.append(scored)
        
        if verbose:
            if scored.get("error"):
                print(f"❌ {scored['error']}")
            else:
                if source == "cache":
                    print(
                        f"📦 Score: {scored['total_score']}/{scored.get('score_out_of', 100)} "
                        f"{scored['grade']} | Coverage: {scored.get('data_coverage_pct', 0)}% (cache)"
                    )
                elif source == "stale-cache":
                    print(
                        f"📦 Score: {scored['total_score']}/{scored.get('score_out_of', 100)} "
                        f"{scored['grade']} | Coverage: {scored.get('data_coverage_pct', 0)}% (stale cache)"
                    )
                else:
                    print(
                        f"✅ Score: {scored['total_score']}/{scored.get('score_out_of', 100)} "
                        f"{scored['grade']} | Coverage: {scored.get('data_coverage_pct', 0)}%"
                    )
                    if data.get("moneycontrolFallbackFailed"):
                        failed_fields = data.get("moneycontrolFailedFields", "") or "unknown"
                        fallback_error = data.get("moneycontrolFallbackError", "") or "parse_or_field_missing"
                        print(
                            f"     ⚠️ Moneycontrol fallback failed for {symbol} "
                            f"(fields: {failed_fields}; reason: {fallback_error})"
                        )
                    if data.get("screenerFallbackFailed"):
                        screener_failed_fields = data.get("screenerFailedFields", "") or "unknown"
                        screener_error = data.get("screenerFallbackError", "") or "parse_or_field_missing"
                        print(
                            f"     ⚠️ Screener fallback failed for {symbol} "
                            f"(fields: {screener_failed_fields}; reason: {screener_error})"
                        )
        
        time.sleep(SLEEP_BETWEEN_FETCHES)

    if use_cache and cache_dirty:
        save_data_cache(cache, cache_file)
    
    df = pd.DataFrame([r for r in results if not r.get("error")])
    
    if df.empty:
        print("\n⚠️ No stocks returned data. Check symbols and network.")
        return df
    
    # Drop detail column for the summary view
    if "details" in df.columns:
        df = df.drop(columns=["details"])
    
    df = df.sort_values("total_score", ascending=False).reset_index(drop=True)
    df.index = df.index + 1  # 1-based ranking
    df.index.name = "Rank"
    
    return df


def print_results(df: pd.DataFrame):
    """Pretty-print the results table."""
    if df.empty:
        return
    
    print("\n" + "=" * 100)
    print(f"  SWING TRADE CANDIDATES — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 100)
    
    display_cols = [
        "symbol", "total_score", "score_out_of", "data_coverage_pct", "grade",
        "intrinsicScore", "dataConfidence", "finalScore",
        "profitability", "balance_sheet", "valuation", "quality", "technicals",
        "pe", "roe", "de", "rsi", "red_flags",
    ]
    
    existing_cols = [c for c in display_cols if c in df.columns]
    
    # Column headers
    headers = {
        "symbol": "Symbol", "total_score": "Score", "score_out_of": "OutOf", "data_coverage_pct": "Coverage%", "grade": "Grade",
        "intrinsicScore": "Intrinsic", "dataConfidence": "Confidence%", "finalScore": "Final",
        "profitability": "Prof/30", "balance_sheet": "BS/20", "valuation": "Val/25",
        "quality": "Qual/15", "technicals": "Tech/10",
        "pe": "P/E", "roe": "ROE%", "de": "D/E", "rsi": "RSI", "red_flags": "Flags",
    }
    
    print(df[existing_cols].rename(columns=headers).to_string())
    
    # Summary stats
    print("\n" + "-" * 60)
    strong = len(df[df["total_score"] >= 80])
    buy = len(df[(df["total_score"] >= 70) & (df["total_score"] < 80)])
    watch = len(df[(df["total_score"] >= 60) & (df["total_score"] < 70)])
    print(f"  🟢 Strong Buy: {strong}  |  🟢 Buy: {buy}  |  🟡 Watchlist: {watch}")
    print(f"  Total screened: {len(df)}  |  Avg score: {df['total_score'].mean():.1f}")
    print("-" * 60)


def export_to_excel(df: pd.DataFrame, filename: str = "swing_candidates.xlsx"):
    """Export results to a color-coded Excel file."""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            # Remove details column if present
            export_df = df.drop(columns=["details"], errors="ignore")
            export_df.to_excel(writer, sheet_name="Rankings", index=True)
            
            ws = writer.sheets["Rankings"]
            
            # Color-code the score column
            score_col = None
            for col_idx, col in enumerate(export_df.columns, start=2):  # +2 for index col
                if col == "total_score":
                    score_col = col_idx
                    break
            
            if score_col:
                for row_idx in range(2, len(export_df) + 2):
                    cell = ws.cell(row=row_idx, column=score_col)
                    try:
                        val = float(cell.value)
                        if val >= 80:
                            cell.fill = PatternFill("solid", fgColor="00C853")
                        elif val >= 70:
                            cell.fill = PatternFill("solid", fgColor="64DD17")
                        elif val >= 60:
                            cell.fill = PatternFill("solid", fgColor="FFD600")
                        elif val >= 50:
                            cell.fill = PatternFill("solid", fgColor="FF9100")
                        else:
                            cell.fill = PatternFill("solid", fgColor="FF1744")
                    except:
                        pass
            
            # Auto-width columns
            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
        
        print(f"\n📊 Exported to {filename}")
    except ImportError:
        # Fallback: CSV
        csv_name = filename.replace(".xlsx", ".csv")
        df.to_csv(csv_name)
        print(f"\n📊 Exported to {csv_name} (install openpyxl for Excel format)")


# ─────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Indian Stock Swing Trading Screener")
    parser.add_argument("--symbols", nargs="+", help="Space-separated stock symbols (e.g., TCS.NS INFY.NS)")
    parser.add_argument("--csv", help="CSV file with a 'Symbol' column")
    parser.add_argument("--export", default="Nifty50Candidates.xlsx", help="Output Excel filename")
    parser.add_argument("--top", type=int, default=100, help="Show top N results")
    parser.add_argument("--quiet", action="store_true", help="Minimal output")
    parser.add_argument("--cache-file", default="yfinance_cache.xlsx", help="Excel cache file for fetched market data")
    parser.add_argument("--cache-hours", type=float, default=24, help="Reuse cache entries newer than this many hours")
    parser.add_argument("--refresh-cache", action="store_true", help="Ignore cache and refresh all symbols from yfinance")
    parser.add_argument("--no-cache", action="store_true", help="Disable cache and always fetch from yfinance")
    parser.add_argument(
        "--cache",
        choices=["auto", "on", "off"],
        default="off",
        help="Cache mode: auto=fresh cache only, on=use cache whenever available, off=disable cache",
    )
    args = parser.parse_args()

    cache_mode = args.cache
    if args.no_cache:
        cache_mode = "off"
    
    # Determine symbol list
    if args.symbols:
        symbols = args.symbols
    elif args.csv:
        csv_df = pd.read_csv(args.csv)
        col = [c for c in csv_df.columns if "symbol" in c.lower()][0]
        symbols = csv_df[col].tolist()
        # Add .NS suffix if missing
        symbols = [s if ".NS" in s or ".BO" in s else f"{s}.NS" for s in symbols]
    else:
        symbols = DEFAULT_SYMBOLS
    
    print(f"\n🔍 Screening {len(symbols)} stocks through Buffett/Lynch/Graham rules...\n")
    
    df = run_screener(
        symbols,
        verbose=not args.quiet,
        use_cache=cache_mode != "off",
        cache_mode=cache_mode,
        cache_file=args.cache_file,
        cache_max_age_hours=args.cache_hours,
        refresh_cache=args.refresh_cache,
    )
    
    if not df.empty:
        print_results(df.head(args.top))
        export_to_excel(df, args.export)


if __name__ == "__main__":
    main()
