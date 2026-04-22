import pandas as pd
from datetime import datetime


def printResults(df: pd.DataFrame):
    if df is None or df.empty:
        return

    print("\n" + "=" * 100)
    print(f"  SWING TRADE CANDIDATES — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 100)

    displayCols = [
        "symbol",
        "totalScore",
        "scoreOutOf",
        "dataCoveragePct",
        "grade",
        "intrinsicScore",
        "dataConfidence",
        "finalScore",
        "profitability",
        "balanceSheetScore",
        "valuation",
        "quality",
        "technicals",
        "marketCapCr",
        "peRatio",
        "roe",
        "debtToEquity",
        "rsi14",
        "flags",
    ]
    availableCols = [col for col in displayCols if col in df.columns]
    displayDf = df[availableCols].copy()
    displayDf = displayDf.rename(
        columns={
            "symbol": "Symbol",
            "totalScore": "Score",
            "scoreOutOf": "OutOf",
            "dataCoveragePct": "Coverage%",
            "grade": "Grade",
            "intrinsicScore": "Intrinsic",
            "dataConfidence": "Confidence%",
            "finalScore": "Final",
            "profitability": "Prof/30",
            "balanceSheetScore": "BS/20",
            "valuation": "Val/25",
            "quality": "Qual/15",
            "technicals": "Tech/10",
            "marketCapCr": "MCapCr",
            "peRatio": "PE",
            "roe": "ROE",
            "debtToEquity": "D/E",
            "rsi14": "RSI",
            "flags": "Flags",
        }
    )
    print(displayDf.to_string())

    print("\n" + "-" * 60)
    strong = len(df[df["totalScore"] >= 80])
    buy = len(df[(df["totalScore"] >= 70) & (df["totalScore"] < 80)])
    watch = len(df[(df["totalScore"] >= 60) & (df["totalScore"] < 70)])
    print(f"  🟢 Strong Buy: {strong}  |  🟢 Buy: {buy}  |  🟡 Watchlist: {watch}")
    print(f"  Total screened: {len(df)}  |  Avg score: {df['totalScore'].mean():.1f}")
    print("-" * 60)


def exportToExcel(df: pd.DataFrame, filename: str = "swingCandidates.xlsx"):
    if df is None or df.empty:
        print("\n⚠️ No data to export.")
        return
    try:
        from openpyxl.styles import PatternFill

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            exportCols = [
                "symbol",
                "name",
                "sector",
                "industry",
                "totalScore",
                "scoreOutOf",
                "dataCoveragePct",
                "grade",
                "intrinsicScore",
                "dataConfidence",
                "finalScore",
                "profitabilityScore",
                "balanceSheetScore",
                "valuationScore",
                "qualityScore",
                "technicalScore",
                "marketCapCr",
                "currentPrice",
                "peRatio",
                "roe",
                "debtToEquity",
                "rsi14",
                "atr",
                "adx",
                "plusDi",
                "minusDi",
                "strongTrend",
                "buySignal",
                "flags",
            ]
            availableExportCols = [column for column in exportCols if column in df.columns]
            exportDf = df[availableExportCols].copy()
            exportDf.to_excel(writer, sheet_name="Rankings", index=True)

            worksheet = writer.sheets["Rankings"]
            scoreColumn = None
            for colIndex, columnName in enumerate(exportDf.columns, start=2):
                if columnName == "totalScore":
                    scoreColumn = colIndex
                    break

            if scoreColumn:
                for rowIndex in range(2, len(exportDf) + 2):
                    cell = worksheet.cell(row=rowIndex, column=scoreColumn)
                    try:
                        value = float(cell.value)
                        if value >= 80:
                            cell.fill = PatternFill("solid", fgColor="00C853")
                        elif value >= 70:
                            cell.fill = PatternFill("solid", fgColor="64DD17")
                        elif value >= 60:
                            cell.fill = PatternFill("solid", fgColor="FFD600")
                        elif value >= 50:
                            cell.fill = PatternFill("solid", fgColor="FF9100")
                        else:
                            cell.fill = PatternFill("solid", fgColor="FF1744")
                    except Exception:
                        pass

            for column in worksheet.columns:
                maxLen = max(len(str(cell.value or "")) for cell in column)
                worksheet.column_dimensions[column[0].column_letter].width = min(maxLen + 2, 30)

        print(f"\n📊 Exported to {filename}")
    except ImportError:
        csvName = filename.replace(".xlsx", ".csv")
        df.to_csv(csvName)
        print(f"\n📊 Exported to {csvName} (install openpyxl for Excel format)")
